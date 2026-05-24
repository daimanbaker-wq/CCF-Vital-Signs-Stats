"""
Impact Cornwall — Vital Stats Catalogue Validator
==================================================
Usage:
    python3 validate_vital_stats.py

Expects "Partially completed.xlsx" in the same directory.
Outputs "Vital_Stats_Validated.xlsx" with:
  - Sheet 1: original data + populated "Correct?" column
  - Sheet 2: audit trail (resolved URL, facts searched, what was found)

Requirements:
    pip install openpyxl requests beautifulsoup4 lxml
"""

import re
import time
import requests
import openpyxl
from openpyxl.styles import PatternFill, Font
from bs4 import BeautifulSoup
from urllib.parse import urlparse
from pathlib import Path

# ── Config ────────────────────────────────────────────────────────────────────

INPUT_FILE  = "Partially completed.xlsx"
OUTPUT_FILE = "Vital_Stats_Validated.xlsx"
REQUEST_TIMEOUT = 15       # seconds per HTTP request
RATE_LIMIT_DELAY = 1.0     # seconds between requests (be polite)
MAX_RETRIES = 2

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (compatible; ImpactCornwallValidator/1.0; "
        "+https://impactcornwall.org.uk)"
    )
}

# Colours for "Correct?" column
GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GREY   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

# ── Source agency mapping (for missing URLs) ──────────────────────────────────

SOURCE_MAP = [
    # (keyword list, agency name, candidate URL stem)
    (["population", "resident", "inhabitants", "census"],
     "ONS Census",
     "https://www.ons.gov.uk/peoplepopulationandcommunity/populationandmigration"),

    (["unemployment", "claimant", "jobseeker", "job seeker", "employment rate",
      "economic inactivity", "workforce", "claimant count"],
     "Nomis (ONS Labour Market)",
     "https://www.nomisweb.co.uk/"),

    (["deprivation", "imd", "index of multiple deprivation", "deprived"],
     "MHCLG Deprivation Data",
     "https://www.gov.uk/government/statistics/english-indices-of-deprivation-2019"),

    (["health", "life expectancy", "obesity", "nhs", "hospital", "gp",
      "mental health", "mortality", "diabetes", "cancer"],
     "NHS Digital / OHID",
     "https://fingertips.phe.org.uk/profile/public-health-outcomes-framework"),

    (["school", "education", "pupil", "ofsted", "gcse", "attainment",
      "exclusion", "special educational needs", "sen"],
     "DfE / Ofsted",
     "https://www.gov.uk/government/organisations/department-for-education"),

    (["earnings", "wage", "salary", "median pay", "income", "gdp",
      "gross value added", "gva"],
     "ONS Earnings / HMRC",
     "https://www.ons.gov.uk/employmentandlabourmarket/peopleinwork/earningsandworkinghours"),

    (["housing", "house price", "affordability", "rent", "social housing",
      "homelessness"],
     "ONS Housing / DLUHC",
     "https://www.ons.gov.uk/peoplepopulationandcommunity/housing"),

    (["broadband", "connectivity", "digital", "internet access"],
     "Ofcom Connected Nations",
     "https://www.ofcom.org.uk/research-and-data/telecoms-research/connected-nations"),

    (["transport", "road", "rail", "bus", "commute", "journey time"],
     "DfT Transport Statistics",
     "https://www.gov.uk/government/organisations/department-for-transport"),

    (["crime", "police", "offence", "anti-social"],
     "Home Office Crime Stats",
     "https://www.gov.uk/government/collections/crime-statistics"),

    (["business", "enterprise", "self-employed", "sme", "startup", "company"],
     "ONS Business Register",
     "https://www.ons.gov.uk/businessindustryandtrade/business/activitysizeandlocation"),

    (["tourism", "visitor", "holiday"],
     "Visit Cornwall / VisitBritain",
     "https://www.visitcornwall.com/industry/research-statistics"),

    (["poverty", "child poverty", "fuel poverty", "food bank"],
     "DWP / Joseph Rowntree Foundation",
     "https://www.gov.uk/government/organisations/department-for-work-pensions"),
]


# ── URL utilities ─────────────────────────────────────────────────────────────

def resolve_url(raw_url: str, session: requests.Session) -> tuple[str, str]:
    """
    Follow redirects and return (final_url, status).
    status: 'ok' | 'timeout' | 'error:<code>' | 'exception:<msg>'
    """
    raw_url = raw_url.strip()
    if not raw_url.startswith("http"):
        raw_url = "https://" + raw_url

    for attempt in range(MAX_RETRIES):
        try:
            r = session.head(
                raw_url, allow_redirects=True,
                timeout=REQUEST_TIMEOUT, headers=HEADERS
            )
            return r.url, f"ok:{r.status_code}"
        except requests.Timeout:
            if attempt == MAX_RETRIES - 1:
                return raw_url, "timeout"
        except requests.RequestException as e:
            if attempt == MAX_RETRIES - 1:
                return raw_url, f"exception:{str(e)[:80]}"
        time.sleep(0.5)
    return raw_url, "unknown"


def fetch_page_text(url: str, session: requests.Session) -> tuple[str, str]:
    """
    Fetch page and return (text_content, status).
    Falls back to GET if HEAD succeeded but GET body is needed.
    """
    try:
        r = session.get(
            url, timeout=REQUEST_TIMEOUT, headers=HEADERS,
            allow_redirects=True
        )
        r.raise_for_status()

        content_type = r.headers.get("Content-Type", "")
        if "html" in content_type:
            soup = BeautifulSoup(r.text, "lxml")
            # Strip scripts/styles
            for tag in soup(["script", "style", "nav", "footer", "header"]):
                tag.decompose()
            return soup.get_text(separator=" ", strip=True), "ok"
        elif "pdf" in content_type:
            return "", "pdf:not-parsed"
        else:
            return r.text, "ok"

    except requests.HTTPError as e:
        return "", f"http:{e.response.status_code}"
    except requests.Timeout:
        return "", "timeout"
    except requests.RequestException as e:
        return "", f"exception:{str(e)[:80]}"


# ── Fact extraction ───────────────────────────────────────────────────────────

# Patterns for verifiable data points
FACT_PATTERNS = [
    # Percentages  e.g. "23.4%", "23.4 per cent"
    re.compile(r'\b(\d[\d,]*(?:\.\d+)?)\s*(?:%|per\s+cent)\b', re.IGNORECASE),
    # Integers / large numbers  e.g. "578,324"  "£12.3 million"
    re.compile(r'£?\b(\d[\d,]*(?:\.\d+)?)\s*(?:million|billion|thousand|k\b)?', re.IGNORECASE),
    # Years  e.g. "in 2023"
    re.compile(r'\b(20\d{2}|19\d{2})\b'),
    # Ranks / ordinals  e.g. "13th most deprived"
    re.compile(r'\b(\d+(?:st|nd|rd|th))\b', re.IGNORECASE),
    # Ratios / rates  e.g. "1 in 4"
    re.compile(r'\b(1\s+in\s+\d+)\b', re.IGNORECASE),
    # Named places (Cornwall already implied; catch specific towns/areas)
    re.compile(r'\b(Cornwall|Penwith|Camborne|Redruth|Truro|Penzance|Falmouth|Newquay|'
               r'Bodmin|Helston|St\s+Austell|Hayle|Launceston|St\s+Ives)\b', re.IGNORECASE),
]


def extract_facts(sentence: str) -> list[str]:
    """Return list of distinct verifiable tokens from sentence."""
    facts = set()
    for pattern in FACT_PATTERNS:
        for m in pattern.finditer(sentence):
            token = m.group(0).strip().lower()
            # Skip trivial years/numbers that are too common
            if len(token) >= 2:
                facts.add(token)
    return list(facts)


def search_facts_in_text(facts: list[str], page_text: str) -> tuple[list[str], list[str]]:
    """
    Returns (found_facts, missing_facts).
    Simple case-insensitive substring search.
    """
    text_lower = page_text.lower()
    # Normalise numbers: remove commas for matching
    text_normalised = text_lower.replace(",", "")

    found, missing = [], []
    for fact in facts:
        fact_norm = fact.replace(",", "")
        if fact_norm in text_normalised or fact in text_lower:
            found.append(fact)
        else:
            missing.append(fact)
    return found, missing


# ── URL parsing from cell ─────────────────────────────────────────────────────

def parse_url_cell(cell_value: str) -> list[dict]:
    """
    Parse a URL cell which may contain:
      - A single URL
      - Multiple URLs separated by newlines, semicolons, or spaces
      - Inline annotation like: "URL1 [supports claim A]; URL2 [supports claim B]"
    Returns list of {"url": str, "note": str}
    """
    if not cell_value or str(cell_value).strip().lower() in ("not available", "n/a", ""):
        return []

    raw = str(cell_value).strip()
    entries = []

    # Split on newlines first, then semicolons
    for chunk in re.split(r'[\n;]+', raw):
        chunk = chunk.strip()
        if not chunk:
            continue

        # Extract optional inline note in brackets
        note_match = re.search(r'\[([^\]]+)\]', chunk)
        note = note_match.group(1).strip() if note_match else ""

        # Extract URL
        url_match = re.search(r'https?://\S+', chunk)
        if url_match:
            url = url_match.group(0).rstrip(".,;)")
            entries.append({"url": url, "note": note})
        elif re.match(r'https?://', chunk):
            # Full chunk is URL
            entries.append({"url": chunk.split()[0], "note": note})

    return entries


# ── Source suggestion ─────────────────────────────────────────────────────────

def suggest_source(sentence: str) -> tuple[str, str]:
    """Return (agency_name, candidate_url) based on sentence keywords."""
    s = sentence.lower()
    for keywords, agency, url in SOURCE_MAP:
        if any(kw in s for kw in keywords):
            return agency, url
    return "Unknown — manual review needed", ""


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    input_path = Path(INPUT_FILE)
    if not input_path.exists():
        print(f"ERROR: '{INPUT_FILE}' not found in current directory.")
        print(f"Current directory: {Path.cwd()}")
        return

    print(f"Loading {INPUT_FILE}...")
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    # ── Detect header row ──────────────────────────────────────────────────────
    header_row = None
    col_sentence = col_url = col_correct = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        for col_idx, val in enumerate(row, start=1):
            if val is None:
                continue
            v = str(val).lower().strip()
            if "sentence" in v or "claim" in v or "fact" in v:
                col_sentence = col_idx
            elif "url" in v or "source" in v or "link" in v:
                col_url = col_idx
            elif "correct" in v or "valid" in v or "check" in v:
                col_correct = col_idx
        if col_sentence and col_url:
            header_row = row_idx
            break

    if not header_row:
        print("ERROR: Could not detect header row. Attempting column A=Sentence, B=URL, C=Correct?")
        header_row = 1
        col_sentence, col_url, col_correct = 1, 2, 3

    # Create "Correct?" column if missing
    if col_correct is None:
        col_correct = (col_url or 2) + 1
        ws.cell(row=header_row, column=col_correct, value="Correct?")
        print(f"  Added 'Correct?' column at column {col_correct}")

    print(f"  Headers at row {header_row}: Sentence=col{col_sentence}, "
          f"URL=col{col_url}, Correct?=col{col_correct}")

    # ── Audit sheet ────────────────────────────────────────────────────────────
    if "Audit" in wb.sheetnames:
        del wb["Audit"]
    ws_audit = wb.create_sheet("Audit")
    audit_headers = [
        "Row", "Sentence (truncated)", "Raw URL", "Resolved URL",
        "URL Status", "Facts Extracted", "Facts Found", "Facts Missing",
        "Verdict", "Suggested Agency", "Suggested URL"
    ]
    ws_audit.append(audit_headers)
    for cell in ws_audit[1]:
        cell.font = Font(bold=True)

    # ── Process rows ───────────────────────────────────────────────────────────
    session = requests.Session()
    session.headers.update(HEADERS)

    data_rows = list(ws.iter_rows(min_row=header_row + 1))
    total = len(data_rows)
    print(f"\nProcessing {total} data rows...\n")

    for i, row in enumerate(data_rows, start=1):
        row_num = header_row + i

        sentence_cell = row[col_sentence - 1] if col_sentence <= len(row) else None
        url_cell      = row[col_url - 1]      if col_url <= len(row) else None
        correct_cell  = ws.cell(row=row_num, column=col_correct)

        sentence = str(sentence_cell.value).strip() if sentence_cell and sentence_cell.value else ""
        url_raw  = str(url_cell.value).strip()      if url_cell and url_cell.value else ""

        print(f"[{i}/{total}] {sentence[:80]}...")

        # Facts to search for
        facts = extract_facts(sentence)

        # ── Case 1: No URL ────────────────────────────────────────────────────
        if not url_raw or url_raw.lower() in ("not available", "n/a", ""):
            agency, candidate_url = suggest_source(sentence)
            correct_cell.value = "? (no URL)"
            correct_cell.fill = GREY
            ws_audit.append([
                row_num,
                sentence[:120],
                "(none)",
                "",
                "no_url",
                "; ".join(facts),
                "",
                "",
                "? (no URL)",
                agency,
                candidate_url
            ])
            continue

        # ── Case 2: Validate each URL ─────────────────────────────────────────
        url_entries = parse_url_cell(url_raw)
        if not url_entries:
            # Raw value present but couldn't parse URL
            correct_cell.value = "? (unparseable URL)"
            correct_cell.fill = YELLOW
            ws_audit.append([row_num, sentence[:120], url_raw, "", "parse_failed",
                             "; ".join(facts), "", "", "? (unparseable URL)", "", ""])
            continue

        verdicts = []

        for entry in url_entries:
            raw_url = entry["url"]
            note    = entry["note"]

            time.sleep(RATE_LIMIT_DELAY)

            # Resolve redirect
            resolved_url, url_status = resolve_url(raw_url, session)

            if "timeout" in url_status or "exception" in url_status:
                verdicts.append("?")
                ws_audit.append([
                    row_num, sentence[:120], raw_url, resolved_url, url_status,
                    "; ".join(facts), "", "", "? (unreachable)", "", ""
                ])
                continue

            # Fetch page content
            page_text, fetch_status = fetch_page_text(resolved_url, session)

            if not page_text:
                verdict = "? (JS-rendered or fetch failed)"
                ws_audit.append([
                    row_num, sentence[:120], raw_url, resolved_url,
                    fetch_status, "; ".join(facts), "", "", verdict, "", ""
                ])
                verdicts.append("?")
                continue

            # Search for facts
            found, missing = search_facts_in_text(facts, page_text)

            # Verdict logic:
            # ✓ if >50% of numeric/specific facts found
            # ✗ if facts extracted but none found
            # ? if no verifiable facts could be extracted
            if not facts:
                verdict = "? (no verifiable facts extracted)"
            elif len(found) / len(facts) >= 0.5:
                verdict = "✓"
            elif found:
                verdict = "✓ (partial)"
            else:
                verdict = "✗"

            ws_audit.append([
                row_num, sentence[:120], raw_url, resolved_url,
                fetch_status,
                "; ".join(facts),
                "; ".join(found),
                "; ".join(missing),
                verdict,
                note,
                ""
            ])
            verdicts.append(verdict)

        # Consolidate verdict across multiple URLs for this row
        if all(v == "✓" for v in verdicts):
            final = "✓"
            fill  = GREEN
        elif any(v == "✓" for v in verdicts):
            final = "✓ (partial)"
            fill  = GREEN
        elif all(v == "✗" for v in verdicts):
            final = "✗"
            fill  = RED
        else:
            final = "?"
            fill  = YELLOW

        correct_cell.value = final
        correct_cell.fill  = fill

    # ── Auto-fit audit columns ─────────────────────────────────────────────────
    for col in ws_audit.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws_audit.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    # ── Save ───────────────────────────────────────────────────────────────────
    out_path = Path(OUTPUT_FILE)
    wb.save(out_path)
    print(f"\n✓ Saved: {out_path.resolve()}")

    # ── Summary ────────────────────────────────────────────────────────────────
    results = [ws.cell(row=header_row + i + 1, column=col_correct).value
               for i in range(total)]
    confirmed  = sum(1 for v in results if v and "✓" in str(v))
    failed     = sum(1 for v in results if v == "✗")
    uncertain  = sum(1 for v in results if v and "?" in str(v))
    no_url     = sum(1 for v in results if v and "no URL" in str(v))

    print(f"\n{'─'*50}")
    print(f"  ✓  Confirmed:       {confirmed}")
    print(f"  ✗  Not found:       {failed}")
    print(f"  ?  Uncertain:       {uncertain}")
    print(f"  ○  No URL (sourced): {no_url}")
    print(f"{'─'*50}")
    print(f"  See 'Audit' tab for full detail on each row.")


if __name__ == "__main__":
    main()
