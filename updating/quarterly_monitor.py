#!/usr/bin/env python3
"""CCF Vital Stats Database — Quarterly Monitor

Validates statistical claims against their source URLs using Playwright.
Flags LINK_BROKEN (unreachable URLs) and DATA_DRIFT (stat missing/changed).

Usage:
    python3 updating/quarterly_monitor.py                  # run on all rows
    python3 updating/quarterly_monitor.py --dry-run        # report only, no writes
    python3 updating/quarterly_monitor.py --limit 10       # first 10 rows
    python3 updating/quarterly_monitor.py --rerun          # re-check ALL rows (even CONFIRMED)
"""

import json
import re
import sys
import os
import time
from datetime import datetime
from pathlib import Path
from collections import Counter

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from bs4 import BeautifulSoup

# ─── Paths ─────────────────────────────────────────────────────────────────────

HERE = Path(__file__).resolve().parent.parent
DATA_DIR = HERE / "data"
CONFIG_DIR = HERE / "config"
REPORTS_DIR = HERE / "reports"
XLSX_IN = HERE / "Vital_Stats_Verified (June 19).xlsx"
XLSX_OUT = HERE / "Vital_Stats_Verified (June 19).xlsx"
REPORT_OUT = REPORTS_DIR / "Quarterly_Exception_Report.md"
CONFIG_FILE = CONFIG_DIR / "maintenance_config.json"

# ═════════════════════════════════════════════════════════════════════════════════
# Helpers
# ═════════════════════════════════════════════════════════════════════════════════

def load_config():
    with open(CONFIG_FILE) as f:
        return json.load(f)


def normalize(text: str) -> str:
    """Normalize text for fuzzy comparison: collapse whitespace, strip."""
    return re.sub(r'\s+', ' ', text).strip()


def is_pdf_url(url: str) -> bool:
    return url.lower().rstrip('/').endswith('.pdf')


def extract_key_numbers(text: str) -> list:
    """Extract notable numbers from a sentence for targeted matching."""
    # Match: currency amounts, percentages, comma-formatted numbers, simple integers
    results = re.findall(r'[£$€]?\d{1,3}(?:,\d{3})*(?:\.\d+)?%?', text)
    # Also match standalone 4+ digit numbers (e.g., 2023, 15000)
    results += re.findall(r'\b\d{4,}\b', text)
    return list(set(results))  # deduplicate


def sentence_keywords(sentence: str) -> str:
    """Strip numbers and common filler, return lowercased keywords."""
    s = sentence.lower()
    s = re.sub(r'[£$€]\d[\d,%. ]+', ' ', s)  # remove money amounts
    s = re.sub(r'\d[\d,%.]*', ' ', s)          # remove numbers
    s = re.sub(r'[^\w\s]', ' ', s)             # remove punctuation
    words = s.split()
    stopwords = {'the', 'a', 'an', 'of', 'in', 'to', 'and', 'is', 'are',
                 'was', 'were', 'has', 'have', 'had', 'been', 'be', 'by',
                 'at', 'from', 'for', 'on', 'as', 'with', 'that', 'its',
                 'this', 'which', 'their', 'per', 'cent'}
    return ' '.join(w for w in words if w not in stopwords and len(w) > 2)


def strip_extra_info(sentence: str) -> str:
    """Strip parenthetical citations like (ONS, Cornwall Council)."""
    return re.sub(r'\([^)]*\)', '', sentence).strip()


# ═════════════════════════════════════════════════════════════════════════════════
# Scraping
# ═════════════════════════════════════════════════════════════════════════════════

def fetch_page_text(playwright, url: str, config: dict, timeout: int = 60) -> str:
    """Fetch page content via Playwright, return extracted text."""
    try:
        browser = playwright.chromium.launch(
            headless=True,
            channel='chrome',
            args=['--no-sandbox', '--disable-gpu', '--disable-software-rasterizer']
        )
        ctx = browser.new_context(
            user_agent=config["scraper"]["user_agent"],
            viewport={"width": 1280, "height": 720},
            locale="en-GB",
        )
        page = ctx.new_page()
        page.set_default_timeout(config["scraper"]["timeout_ms"])

        # Block non-essential resources for speed
        page.route(re.compile(r'\.(png|jpg|jpeg|gif|svg|webp|ico|woff|woff2|ttf|eot|mp4|webm|ogg|avi|mov)(\?.*)?$', re.I),
                   lambda route: route.abort())

        # First attempt — fast initial load
        try:
            resp = page.goto(url, wait_until="commit", timeout=timeout * 1000)
        except Exception:
            # Retry with relaxed timeout
            try:
                resp = page.goto(url, wait_until="commit", timeout=(timeout + 30) * 1000)
            except Exception as e:
                browser.close()
                return f"FETCH_ERROR:{type(e).__name__}:Timed out after {timeout+30}s"

        if resp and resp.status >= 400:
            browser.close()
            return f"HTTP_ERROR:{resp.status}"

        # Get the HTML we have (even if DOMContentLoaded hasn't fired)
        try:
            page.wait_for_load_state("domcontentloaded", timeout=15000)
        except Exception:
            pass
        content = page.content()
        browser.close()

        soup = BeautifulSoup(content, "lxml")
        # Remove script/style/nav/footer clutter
        for tag in soup(["script", "style", "nav", "footer", "header", "aside"]):
            tag.decompose()
        text = soup.get_text(separator=" ", strip=True)
        return text or "EMPTY_PAGE"

    except Exception as e:
        return f"FETCH_ERROR:{type(e).__name__}:{str(e)[:120]}"


# ═════════════════════════════════════════════════════════════════════════════════
# Verification
# ═════════════════════════════════════════════════════════════════════════════════

def verify_stat(sentence: str, page_text: str) -> dict:
    """Compare a stat sentence against fetched page text.
    
    Returns: {
        "verdict": "CONFIRMED" | "PARTIAL" | "NOT_FOUND" | "ERROR",
        "confidence": int (0-100),
        "evidence": str (matched snippet)
    }
    """
    if not page_text or page_text.startswith("HTTP_ERROR") or page_text.startswith("FETCH_ERROR"):
        return {"verdict": "UNREACHABLE", "confidence": 0, "evidence": page_text[:200]}

    if page_text == "EMPTY_PAGE":
        return {"verdict": "UNREACHABLE", "confidence": 0, "evidence": "Page returned no readable text"}

    clean_sentence = strip_extra_info(sentence)
    norm_sentence = normalize(clean_sentence.lower())
    norm_text = normalize(page_text.lower())
    text_len = len(norm_text)

    # Strategy 1: Exact match — the full sentence appears in the page
    if norm_sentence in norm_text:
        idx = norm_text.index(norm_sentence)
        snippet = norm_text[max(0, idx - 60):idx + len(norm_sentence) + 60]
        return {"verdict": "CONFIRMED", "confidence": 100, "evidence": snippet[:300]}

    # Strategy 2: Key numbers match in context
    numbers = extract_key_numbers(sentence)
    if numbers:
        # Try to find all numbers in the text
        found_numbers = [n for n in numbers if n.lower() in norm_text]
        if len(found_numbers) >= len(numbers) * 0.5:
            # At least some numbers found — now check keyword overlap
            kw_sentence = sentence_keywords(strip_extra_info(sentence))
            kw_text = sentence_keywords(page_text)

            # Count overlapping keywords
            sentence_words = set(kw_sentence.split())
            if sentence_words:
                overlap = sentence_words & set(kw_text.split())
                overlap_ratio = len(overlap) / len(sentence_words)

                if overlap_ratio >= 0.4 and len(found_numbers) >= len(numbers) * 0.75:
                    # Build evidence from the number match areas
                    evidence_parts = []
                    for n in found_numbers[:3]:
                        idx = norm_text.index(n.lower())
                        evidence_parts.append(
                            norm_text[max(0, idx - 50):idx + len(n) + 50]
                        )
                    evidence = " [...] ".join(p[:200] for p in evidence_parts)
                    confidence = int(60 + (overlap_ratio * 30) + 
                                     (len(found_numbers) / len(numbers) * 10))
                    return {"verdict": "CONFIRMED" if confidence >= 70 else "PARTIAL",
                            "confidence": min(confidence, 98), "evidence": evidence[:300]}
                elif overlap_ratio >= 0.2:
                    return {"verdict": "PARTIAL", "confidence": int(overlap_ratio * 60),
                            "evidence": f"Numbers matched: {found_numbers}. Keyword overlap: {overlap_ratio:.0%}"}

    # Strategy 3: Loose keyword check
    kw_sentence = sentence_keywords(strip_extra_info(sentence))
    kw_text_set = set(sentence_keywords(page_text).split())
    words = kw_sentence.split()
    if words:
        matches = sum(1 for w in words if w in kw_text_set)
        ratio = matches / len(words)
        if ratio >= 0.5:
            return {"verdict": "PARTIAL", "confidence": int(ratio * 60),
                    "evidence": f"Keyword match: {matches}/{len(words)} words ({ratio:.0%})"}
        else:
            return {"verdict": "NOT_FOUND", "confidence": int(ratio * 30),
                    "evidence": f"Only {matches}/{len(words)} keywords matched"}

    return {"verdict": "NOT_FOUND", "confidence": 0, "evidence": "No relevant content found"}


# ═════════════════════════════════════════════════════════════════════════════════
# Excel I/O
# ═════════════════════════════════════════════════════════════════════════════════

def load_dataframe():
    """Load the Excel file as a DataFrame for processing."""
    return pd.read_excel(XLSX_IN, header=1)  # row 2 is header


def write_results(df: pd.DataFrame, stats: dict):
    """Write verified results to a new Excel file with formatting."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Verified Results"

    # Column map: A=None, B=Sentence, C=URL, D=Correct?, E=Source Type, F=Verified URL, G=Verdict, H=Confidence, I=Evidence
    col_map = {
        "Sentence": "B", "URL": "C", "Correct?": "D",
        "Source Type": "E", "Verified URL": "F",
        "Verdict": "G", "Confidence": "H", "Evidence": "I"
    }

    # Colours
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="999999"),
        right=Side(style="thin", color="999999"),
        top=Side(style="thin", color="999999"),
        bottom=Side(style="thin", color="999999"),
    )

    # Write header
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 50
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 70

    # Header row
    for col_letter, col_name in [("B", "Sentence"), ("C", "URL"), ("D", "Correct?"),
                                  ("E", "Source Type"), ("F", "Verified URL"),
                                  ("G", "Verdict"), ("H", "Confidence"), ("I", "Evidence")]:
        cell = ws[f"{col_letter}1"]
        cell.value = col_name
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    # Data rows
    for idx, (_, row) in enumerate(df.iterrows(), start=2):
        ws[f"B{idx}"] = row.get("Sentence", "")
        ws[f"C{idx}"] = row.get("URL", "")
        ws[f"D{idx}"] = row.get("Correct?", "")
        ws[f"E{idx}"] = row.get("Source Type", "")
        ws[f"F{idx}"] = row.get("Verified URL", "")
        ws[f"G{idx}"] = row.get("Verdict", "")
        ws[f"H{idx}"] = row.get("Confidence", 0)
        ws[f"I{idx}"] = str(row.get("Evidence", ""))[:32767]  # Excel cell char limit

        # Conditional formatting
        verdict = str(row.get("Verdict", ""))
        if verdict == "CONFIRMED":
            fill = green_fill
        elif verdict in ("NOT_FOUND", "LINK_BROKEN"):
            fill = red_fill
        elif verdict in ("PARTIAL", "DATA_DRIFT"):
            fill = yellow_fill
        else:
            fill = None

        for col in ["B", "C", "D", "E", "F", "G", "H", "I"]:
            cell = ws[f"{col}{idx}"]
            if fill:
                cell.fill = fill
            cell.border = thin_border

    wb.save(XLSX_OUT)
    print(f"  ✓ Written: {XLSX_OUT}")

    # ── Exception Report ────────────────────────────────────────────────────────
    lines = []
    lines.append(f"# Quarterly Exception Report — {datetime.now().strftime('%d %B %Y')}")
    lines.append("")
    lines.append(f"**Period:** {datetime.now().strftime('%B %Y')}")
    lines.append(f"**Total rows processed:** {stats['total']}")
    lines.append(f"**Run duration:** {stats.get('duration', 'N/A')}")
    lines.append("")
    lines.append("## Summary")
    lines.append("")
    lines.append(f"| Status | Count |")
    lines.append(f"| :--- | ---: |")
    for verdict, count in sorted(stats.get('verdicts', {}).items()):
        lines.append(f"| {verdict} | {count} |")
    lines.append("")

    # Error rows
    errors = stats.get('errors', [])
    if errors:
        lines.append("## Exceptions Requiring Attention")
        lines.append("")
        for err in errors:
            lines.append(f"### {err['sentence'][:100]}")
            lines.append("")
            lines.append(f"- **Verdict:** {err['verdict']}")
            lines.append(f"- **URL:** {err['url']}")
            lines.append(f"- **Evidence:** {err['evidence'][:200]}")
            lines.append("")

    if not errors:
        lines.append("**No exceptions — all stats validated successfully.**")
        lines.append("")

    lines.append("---")
    lines.append(f"*Generated by quarterly_monitor.py on {datetime.now().strftime('%Y-%m-%d %H:%M')}*")
    lines.append("")

    REPORT_OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(REPORT_OUT, "w") as f:
        f.write("\n".join(lines))
    print(f"  ✓ Written: {REPORT_OUT}")


# ═════════════════════════════════════════════════════════════════════════════════
# Main
# ═════════════════════════════════════════════════════════════════════════════════

def main():
    import argparse
    parser = argparse.ArgumentParser(description="CCF Vital Stats Quarterly Monitor")
    parser.add_argument("--dry-run", action="store_true", help="Report only, no writes")
    parser.add_argument("--limit", type=int, default=0, help="Process only first N rows")
    parser.add_argument("--rerun", action="store_true", help="Re-check all rows, even CONFIRMED")
    args = parser.parse_args()

    config = load_config()
    print(f"📋 CCF Vital Stats Quarterly Monitor")
    print(f"   Config: {CONFIG_FILE}")
    print(f"   Data:   {XLSX_IN}")
    print(f"   Dry run: {'✓' if args.dry_run else '✗'}")
    print()

    # Load data
    df = load_dataframe()
    total_rows = len(df)
    print(f"   Loaded {total_rows} rows from spreadsheet")
    print()

    # Determine which rows to process
    if args.rerun:
        to_process = df.index.tolist()
        print(f"   Mode: re-run ALL rows (including CONFIRMED)")
    else:
        # Skip CONFIRMED with high confidence
        to_process = df[df["Verdict"] != "CONFIRMED"].index.tolist()
        print(f"   Mode: incremental — {len(to_process)} rows need checking")

    if args.limit and args.limit < len(to_process):
        to_process = to_process[:args.limit]
        print(f"   Limit: first {args.limit} rows")
    print()

    # Process rows
    start_time = time.time()
    stats = {
        "total": len(to_process),
        "verdicts": Counter(),
        "errors": [],
    }

    if not to_process:
        print("   Nothing to process. Use --rerun to re-check all rows.")
        # Still write report with no changes
        if not args.dry_run:
            stats["duration"] = "0s"
            df_updated = df.copy()
            write_results(df_updated, stats)
        return

    from playwright.sync_api import sync_playwright

    with sync_playwright() as pw:
        for i, idx in enumerate(to_process, 1):
            row = df.loc[idx]
            sentence = str(row.get("Sentence", ""))
            url = str(row.get("Verified URL", "") or row.get("URL", ""))
            source_type = str(row.get("Source Type", ""))
            current_verdict = str(row.get("Verdict", ""))

            if not url or url in ("None", "nan", ""):
                df.at[idx, "Verdict"] = "NO_URL"
                df.at[idx, "Confidence"] = "0"
                df.at[idx, "Evidence"] = "No URL provided"
                stats["verdicts"]["NO_URL"] += 1
                continue

            if source_type == "EXTERNAL":
                df.at[idx, "Verdict"] = "EXTERNAL"
                df.at[idx, "Confidence"] = "0"
                df.at[idx, "Evidence"] = "External data source — manual review required"
                stats["verdicts"]["EXTERNAL"] += 1
                continue

            print(f"  [{i}/{len(to_process)}] {sentence[:70]}...")

            # Fetch page
            page_text = fetch_page_text(pw, url, config)

            # Verify
            result = verify_stat(sentence, page_text)

            # Write results
            df.at[idx, "Verdict"] = result["verdict"]
            df.at[idx, "Confidence"] = str(result["confidence"])
            df.at[idx, "Evidence"] = result["evidence"][:300]

            verdict = result["verdict"]
            confidence = result["confidence"]
            stats["verdicts"][verdict] += 1

            if verdict == "UNREACHABLE":
                stats["errors"].append({
                    "sentence": sentence,
                    "url": url,
                    "verdict": "LINK_BROKEN",
                    "evidence": result["evidence"]
                })
            elif verdict in ("NOT_FOUND", "PARTIAL"):
                stats["errors"].append({
                    "sentence": sentence,
                    "url": url,
                    "verdict": "DATA_DRIFT",
                    "evidence": result["evidence"]
                })

            status_icon = {"CONFIRMED": "✓", "PARTIAL": "~", "NOT_FOUND": "✗",
                           "UNREACHABLE": "💀", "NO_URL": "⛔"}.get(verdict, "?")
            print(f"    {status_icon} {verdict} ({confidence}%)")

            # Be nice to servers — small delay between requests
            if i < len(to_process):
                time.sleep(1.5)

    elapsed = time.time() - start_time
    stats["duration"] = f"{elapsed:.0f}s"

    print()
    print("═══ Results ═══")
    for verdict, count in sorted(stats["verdicts"].items()):
        bar = "█" * count if count <= 50 else "█" * 50 + f"({count})"
        print(f"  {verdict:15s} {count:4d}  {bar}")

    print(f"\n  Total errors flagged: {len(stats['errors'])}")
    print(f"  Duration: {stats['duration']}")

    if not args.dry_run:
        print()
        print("📝 Writing results...")
        write_results(df, stats)
        print()
        print(f"✅ Monitor run complete. See:")
        print(f"   Data:   {XLSX_OUT}")
        print(f"   Report: {REPORT_OUT}")
    else:
        print()
        print("🏁 Dry run complete — no files written.")


if __name__ == "__main__":
    main()
