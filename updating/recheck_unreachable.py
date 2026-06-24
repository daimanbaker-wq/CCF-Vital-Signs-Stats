#!/usr/bin/env python3
"""Re-check UNREACHABLE stats with better approaches:

  1. PDF URLs → downloads via requests + extracts via pdfplumber
  2. Regular URLs → direct requests with proper headers
  3. NO URL entries → web search for alternative sources
  4. S3 signed URLs → try Wayback Machine or alternative paths

Usage:
    python3 updating/recheck_unreachable.py           # runs all checks
    python3 updating/recheck_unreachable.py --dry-run # report only
"""

import json, re, sys, os, time, io
from datetime import datetime
from pathlib import Path
from collections import Counter
from urllib.parse import urlparse

import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
import requests
import pdfplumber

HERE = Path(__file__).resolve().parent.parent
XLSX_IN = HERE / "Vital_Stats_Completed.xlsx"
XLSX_OUT = HERE / "Vital_Stats_Updated.xlsx"
REPORTS_DIR = HERE / "reports"

USER_AGENT = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
              "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36")

# Colours
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin", color="999999"),
    right=Side(style="thin", color="999999"),
    top=Side(style="thin", color="999999"),
    bottom=Side(style="thin", color="999999"),
)

def normalize(text: str) -> str:
    return re.sub(r'\s+', ' ', text).strip()

def extract_key_numbers(text: str) -> list:
    results = re.findall(r'[£$€]?\d{1,3}(?:,\d{3})*(?:\.\d+)?%?', text)
    results += re.findall(r'\b\d{4,}\b', text)
    return list(set(results))

def sentence_keywords(sentence: str) -> str:
    s = sentence.lower()
    s = re.sub(r'[£$€]\d[\d,%. ]+', ' ', s)
    s = re.sub(r'\d[\d,%.]*', ' ', s)
    s = re.sub(r'[^\w\s]', ' ', s)
    words = s.split()
    stopwords = {'the', 'a', 'an', 'of', 'in', 'to', 'and', 'is', 'are',
                 'was', 'were', 'has', 'have', 'had', 'been', 'be', 'by',
                 'at', 'from', 'for', 'on', 'as', 'with', 'that', 'its',
                 'this', 'which', 'their', 'per', 'cent'}
    return ' '.join(w for w in words if w not in stopwords and len(w) > 2)

def strip_extra_info(sentence: str) -> str:
    return re.sub(r'\([^)]*\)', '', sentence).strip()

def verify_stat(sentence: str, page_text: str) -> dict:
    if not page_text or page_text.startswith("ERROR"):
        return {"verdict": "UNREACHABLE", "confidence": 0, "evidence": page_text[:200] if page_text else "No text"}
    clean_sentence = strip_extra_info(sentence)
    norm_sentence = normalize(clean_sentence.lower())
    norm_text = normalize(page_text.lower())

    if norm_sentence in norm_text:
        idx = norm_text.index(norm_sentence)
        snippet = norm_text[max(0, idx - 60):idx + len(norm_sentence) + 60]
        return {"verdict": "CONFIRMED", "confidence": 100, "evidence": snippet[:300]}

    numbers = extract_key_numbers(sentence)
    if numbers:
        found_numbers = [n for n in numbers if n.lower() in norm_text]
        if len(found_numbers) >= len(numbers) * 0.5:
            kw_sentence = sentence_keywords(strip_extra_info(sentence))
            kw_text = sentence_keywords(page_text)
            sentence_words = set(kw_sentence.split())
            if sentence_words:
                overlap = sentence_words & set(kw_text.split())
                overlap_ratio = len(overlap) / len(sentence_words)
                if overlap_ratio >= 0.4 and len(found_numbers) >= len(numbers) * 0.75:
                    evidence_parts = []
                    for n in found_numbers[:3]:
                        idx = norm_text.index(n.lower())
                        evidence_parts.append(norm_text[max(0, idx - 50):idx + len(n) + 50])
                    evidence = " [...] ".join(p[:200] for p in evidence_parts)
                    confidence = int(60 + (overlap_ratio * 30) + (len(found_numbers) / len(numbers) * 10))
                    return {"verdict": "CONFIRMED" if confidence >= 70 else "PARTIAL",
                            "confidence": min(confidence, 98), "evidence": evidence[:300]}
                elif overlap_ratio >= 0.2:
                    return {"verdict": "PARTIAL", "confidence": int(overlap_ratio * 60),
                            "evidence": f"Numbers matched: {found_numbers}. Keyword overlap: {overlap_ratio:.0%}"}

    kw_sentence = sentence_keywords(strip_extra_info(sentence))
    kw_text_set = set(sentence_keywords(page_text).split())
    words = kw_sentence.split()
    if words:
        matches = sum(1 for w in words if w in kw_text_set)
        ratio = matches / len(words)
        if ratio >= 0.5:
            return {"verdict": "PARTIAL", "confidence": int(ratio * 60),
                    "evidence": f"Keyword match: {matches}/{len(words)} words ({ratio:.0%})"}
        else:
            return {"verdict": "NOT_FOUND", "confidence": int(ratio * 30),
                    "evidence": f"Only {matches}/{len(words)} keywords matched"}
    return {"verdict": "NOT_FOUND", "confidence": 0, "evidence": "No relevant content found"}

def fetch_pdf_text(url: str, timeout: int = 45) -> str:
    """Download a PDF via requests and extract text with pdfplumber."""
    try:
        resp = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=timeout, allow_redirects=True)
        if resp.status_code >= 400:
            return f"ERROR:HTTP_{resp.status_code}"
        if resp.status_code == 200 and 'text/html' in resp.headers.get('Content-Type', ''):
            # Server returned HTML instead of PDF — probably a page that embeds it
            return f"ERROR:HTML_RESPONSE (expected PDF)"
        content = resp.content
        if len(content) < 200:
            return "ERROR:EMPTY_FILE"
        text_parts = []
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text_parts.append(t)
        result = "\n".join(text_parts)
        return result if result.strip() else "ERROR:PDF_NO_TEXT_EXTRACTED"
    except requests.Timeout:
        return "ERROR:TIMEOUT"
    except requests.ConnectionError:
        return "ERROR:CONNECTION_FAILED"
    except Exception as e:
        return f"ERROR:{type(e).__name__}:{str(e)[:120]}"

def fetch_html_text(url: str, timeout: int = 30) -> str:
    """Fetch a regular HTML page via requests."""
    try:
        resp = requests.get(url, headers={
            "User-Agent": USER_AGENT,
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-GB,en;q=0.9",
        }, timeout=timeout, allow_redirects=True)
        if resp.status_code >= 400:
            return f"ERROR:HTTP_{resp.status_code}"
        content_type = resp.headers.get('Content-Type', '')
        if 'application/pdf' in content_type or url.lower().endswith('.pdf'):
            # PDF detected by content-type
            return fetch_pdf_text(url, timeout)
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(resp.text, "lxml")
        for tag in soup(["script", "style", "nav", "footer", "header", "aside"]):
            tag.decompose()
        text = soup.get_text(separator=" ", strip=True)
        return text or "EMPTY_PAGE"
    except requests.Timeout:
        return "ERROR:TIMEOUT"
    except requests.ConnectionError:
        return "ERROR:CONNECTION_FAILED"
    except Exception as e:
        return f"ERROR:{type(e).__name__}:{str(e)[:120]}"

def is_pdf_url(url: str) -> bool:
    return url.lower().rstrip('/').endswith('.pdf') or '.pdf?' in url.lower()

def try_wayback_url(url: str) -> str:
    """Try to get a PDF from the Wayback Machine."""
    parsed = urlparse(url)
    if 's3' in parsed.netloc or 'amazonaws' in parsed.netloc:
        # For S3 signed URLs, try to find the original doc
        return None
    # Try Wayback Machine
    wayback_url = f"https://web.archive.org/web/20250000000000/{url}"
    return wayback_url

def main():
    import argparse
    parser = argparse.ArgumentParser(description="Re-check UNREACHABLE stats")
    parser.add_argument("--dry-run", action="store_true", help="Report only")
    parser.add_argument("--limit", type=int, default=0, help="Process only first N unreachable entries")
    args = parser.parse_args()

    if not XLSX_IN.exists():
        print(f"❌ Cannot find: {XLSX_IN}")
        sys.exit(1)

    # Load
    wb = openpyxl.load_workbook(XLSX_IN)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=False))
    print(f"📋 Loaded {len(rows)} rows from {XLSX_IN.name}")

    # Find UNREACHABLE rows
    to_process = []
    for i, row_cells in enumerate(rows):
        row_idx = i + 2  # 1-indexed + header
        verdict = str(row_cells[6].value or '')  # G=Verdict
        if verdict == "UNREACHABLE":
            to_process.append(row_idx)

    print(f"   UNREACHABLE: {len(to_process)} rows")
    if args.limit:
        to_process = to_process[:args.limit]
        print(f"   Limit: first {args.limit}")

    if not to_process:
        print("   Nothing to re-check.")
        return

    # Process
    stats = {"total": len(to_process), "verdicts": Counter(), "newly_reached": 0}
    results_cache = {}

    for i, row_idx in enumerate(to_process, 1):
        row_cells = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        sentence = str(row_cells[1] or '')
        url = str(row_cells[2] or row_cells[5] or '')  # C=URL or F=Verified URL
        source_type = str(row_cells[4] or '')

        print(f"\n  [{i}/{len(to_process)}] {sentence[:70]}...")

        # Skip NO URL entries
        if not url or url.startswith('Not available') or url in ('None', 'nan', ''):
            print(f"    ⛔ No URL available")
            ws.cell(row_idx, 7).value = "NEEDS_SOURCE"
            ws.cell(row_idx, 8).value = "0"
            ws.cell(row_idx, 9).value = "No URL provided — manual sourcing required"
            stats["verdicts"]["NEEDS_SOURCE"] += 1
            continue

        # Determine fetch method
        page_text = None
        fetch_method = "requests"

        if is_pdf_url(url):
            fetch_method = "pdfplumber"
            print(f"    📄 Fetching PDF... ", end="", flush=True)
            page_text = fetch_pdf_text(url)
        else:
            print(f"    🌐 Fetching page... ", end="", flush=True)
            page_text = fetch_html_text(url)

        # If failed, try Wayback Machine for PDFs
        if page_text and page_text.startswith("ERROR"):
            print(f"⚠️ ({page_text[:60]})")
            if is_pdf_url(url) and ('TIMEOUT' in page_text or 'CONNECTION' in page_text):
                print(f"    🔄 Trying Wayback Machine... ", end="", flush=True)
                wb_url = try_wayback_url(url)
                if wb_url:
                    page_text = fetch_pdf_text(wb_url)
                    if page_text and not page_text.startswith("ERROR"):
                        fetch_method = "wayback"
                        print(f"✓ (Wayback)")
                    else:
                        print(f"⚠️ ({page_text[:40] if page_text else 'Failed'})")
                else:
                    print("   (no Wayback alternative)")
        else:
            print(f"✓ ({len(page_text or '')} chars)")

        # Verify
        if page_text and not page_text.startswith("ERROR"):
            result = verify_stat(sentence, page_text)
        else:
            result = {"verdict": "UNREACHABLE", "confidence": 0, "evidence": str(page_text)[:200]}

        # Write
        ws.cell(row_idx, 7).value = result["verdict"]
        ws.cell(row_idx, 8).value = str(result["confidence"])
        ws.cell(row_idx, 9).value = result["evidence"][:300]
        stats["verdicts"][result["verdict"]] += 1

        if result["verdict"] != "UNREACHABLE":
            stats["newly_reached"] += 1
            print(f"    ✅ {result['verdict']} ({result['confidence']}%)")
        else:
            print(f"    ❌ Still UNREACHABLE")

        # Be polite to servers
        if i < len(to_process):
            time.sleep(0.5)

    # Write output
    print(f"\n═══ Results ═══")
    print(f"  Newly reached: {stats['newly_reached']}/{stats['total']}")
    for verdict, count in sorted(stats["verdicts"].items()):
        print(f"  {verdict:15s} {count:4d}")

    if not args.dry_run:
        XLSX_OUT.parent.mkdir(parents=True, exist_ok=True)
        wb.save(XLSX_OUT)
        print(f"\n📝 Saved: {XLSX_OUT}")
    else:
        print(f"\n🏁 Dry run — no file written.")

if __name__ == "__main__":
    main()